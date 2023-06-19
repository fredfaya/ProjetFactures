import csv
import os

import openpyxl
from Logs.my_logger import logger


def dict_to_csv(infos: dict, file_path: str) -> None:
    """
    Function that converts a dictionary to a csv file.
    :param infos: dictionary with extracted information
    :param file_path: path to csv file
    :return: None
    """
    logger.info("Saving the informations in a csv file")
    try:
        # Check if the file exists and is not empty
        if os.path.isfile(file_path) and os.path.getsize(file_path) > 0:
            # File exists and is not empty, open it in append mode
            with open(file_path, 'a', newline='') as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=infos.keys())
                writer.writerow(infos)
        else:
            # File does not exist or is empty, create it and write headers
            with open(file_path, 'w', newline='') as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=infos.keys())
                writer.writeheader()
                writer.writerow(infos)
    except FileNotFoundError as f:
        logger.error("Could not find the file to save the results : {}".format(file_path))
        raise f


def dict_to_excel(infos: dict, file_path: str) -> None:
    """
    Function that converts a dictionary to an Excel file.
    :param infos: dictionary with extracted information
    :param file_path: path to Excel file
    :param sheet_name: name of the sheet to write to (default: 'Sheet1')
    :return: None
    """
    logger.info("Saving the informations in an Excel file")
    try:
        # Create a new workbook if the file does not exist
        if not os.path.isfile(file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Write the keys of the dictionary to the first row of the sheet
            for i, key in enumerate(infos.keys()):
                sheet.cell(row=1, column=i + 1).value = key
            # Start writing the values to the second row of the sheet
            row_num = 2
        else:
            # Open the existing workbook and get the active sheet
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            # Find the next empty row to write to
            row_num = sheet.max_row + 1

        # Write the values to the next empty row of the sheet
        for col_num, key in enumerate(infos.keys(), start=1):
            value = infos[key]
            if isinstance(value, list):
                value = ", ".join(value)  # Join the list into a string
            sheet.cell(row=row_num, column=col_num).value = value

        # Save the workbook to the file path
        workbook.save(file_path)
        logger.info("Information saved to Excel file : {}".format(file_path))
    except FileNotFoundError as f:
        logger.error("Could not find the file to save the results : {}".format(file_path))
        raise f
